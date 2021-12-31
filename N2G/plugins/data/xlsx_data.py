"""
XLSX Data Plugin
****************

This plugin loads data from xlsx tables and transforms it in a dictionary
supported by N2G diagram plugins. Using ``from_dict`` method, this plugin 
loads data into diagram plugin.

Guidelines and Limitations
--------------------------

* `openpyxl <https://pypi.org/project/openpyxl/>`_ >= 3.0.0 library need to be installed: ``pip install openpyxl``
* Nodes and links tabs' first row must contain headers
* nodes tab should have at least ``id`` header, other headers should comply with ``from_dict`` method attributes 
  or simply ignored
* links tab should have at least ``source`` and ``target`` headers, other headers should comply with ``from_dict`` 
  method attributes or simply ignored

Sample Usage
------------

Code to invoke ``xlsx_data``::

    from N2G import drawio_diagram
    from N2G import xlsx_data

    drawio_drawing = drawio_diagram()
    
    xlsx_data(
        drawio_drawing, 
        "./Data/nodes_and_links_data.xlsx", 
        node_tabs="nodes", 
        link_tabs="links"
    )
    
    drawio_drawing.layout(algo="kk")
    drawio_drawing.dump_file(filename="diagram.drawio", folder="./Output/")
    
Where ``nodes_and_links_data.xlsx`` content for nodes tab::

    id   pic    label    bottom_label    top_label    description
    r1          r1       core            1.1.1.1      Core Router
    r2          r2       core            2.2.2.2      Core Router
    r3          r3       edge            3.3.3.3      Edge Router
    
for links tab::
    
    source    src_label    label    target    trgt_label    description
    r1        Gi1/1        DF-10Km  r2        Gi3/4         DF link between R1 and R2
    r3        10GE2/1/1    DF-32Km  r2        Ten1/1        DF link between R3 and R2
    
Support available to translate headers to comply with N2G diagram modules
``from_dict`` or ``from_list`` methods through the use of ``node_headers_map`` and 
``link_headers_map``.  For instance consider this table::

    # nodes tab:
    hostname   lo0_ip    bgp_asn
    r1         1.1.1.1   65123
    r2         1.1.1.2   65123
    r3         1.1.1.3   65123
    
    # links tab:
    device:a    interface:a    label    device:b  interface:b
    r1          Gi1/1          DF-10Km  r2        Gi3/4      
    r3          10GE2/1/1      DF-32Km  r2        Ten1/1     

If ``node_headers_map`` is::
    
    node_headers_map = {
        "id": ["device", "hostname"],
        "tob_label": ["lo0_ip"],
        "bottom_label": ["bgp_asn"]
    }
        
And ``link_headers_map`` is::

    link_headers_map = {
        "source": ["device:a", "hostname:a"], 
        "target": ["device:b", "hostname:b"],
        "src_label": ["interface:a", "ip:a"],
        "trgt_label": ["interface:b", "ip:b"]
    }
    
Above table will be transformed to::

    # nodes tab:
    id         tob_label bottom_label
    r1         1.1.1.1   65123
    r2         1.1.1.2   65123
    r3         1.1.1.3   65123
    
    # links tab:
    source      src_label      label    target    trgt_label
    r1          Gi1/1          DF-10Km  r2        Gi3/4      
    r3          10GE2/1/1      DF-32Km  r2        Ten1/1   
    
API Reference
-------------

.. autofunction:: N2G.plugins.data.xlsx_data.xlsx_data
"""

import logging
import traceback

# initiate logging
log = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    log.error("Failed to import openpyxl module")


def translate_headers(headers, translate_dict):
    """
    Helper function to perform headers translation.

    :param headers: (list) list of headers to translate
    :param translate_dict: (dict) map of headers to translate to

    For example if ``headers`` are::

        headers = ["device:a", "interface:a", "label", "device:b", "interface:b"]

    and ``translate_dict`` is::

        translate_dict = {
            "source": ["device:a", "hostname:a"],
            "target": ["device:b", "hostname:b"],
            "src_label": ["src_label", "interface:a", "ip:a"],
            "trgt_label": ["trgt_label", "interface:b", "ip:b"]
        }

    this function will transform ``headers`` list in::

        headers = ["source", "src_label", "label", "target", "trgt_label"]
    """
    for index, header in enumerate(headers):
        if header in translate_dict:
            continue
        for translate_header, headers_list in translate_dict.items():
            if header in headers_list:
                headers[index] = translate_header
                break


def xlsx_data(
    drawing,
    data,
    node_tabs=["nodes"],
    link_tabs=["links"],
    node_headers_map={"id": ["device", "hostname"]},
    link_headers_map={
        "source": ["device:a", "hostname:a"],
        "target": ["device:b", "hostname:b"],
        "src_label": ["interface:a", "ip:a"],
        "trgt_label": ["interface:b", "ip:b"],
    },
):
    """
    Function to load data from XLSX file and add it to diagram using
    ``from_dict`` method.

    :param drawing: N2G drawing module object
    :param data: (str) OS path to xlsx file to load
    :param node_tabs: (list) list of tabs with nodes data, default ``["nodes"]``
    :param link_tabs: (list) list of tabs with links data, default ``["links"]``
    :param node_headers_map: (dict) dictionary to use to translate node tabs headers
    :param link_headers_map: (dict) dictionary to use to translate link tabs headers
    :return: ``True`` on success and ``False`` on failure to load data
    """
    wb = load_workbook(data, data_only=True, read_only=True)

    graph_dict = {"nodes": [], "links": []}

    try:
        # form graph dictionary
        for node_tab in node_tabs:
            node_headers = []
            for row in wb[node_tab].iter_rows(values_only=True):
                if not node_headers:
                    node_headers = [str(i).strip() for i in row]
                    translate_headers(node_headers, node_headers_map)
                else:
                    graph_dict["nodes"].append(
                        dict(zip(node_headers, [c if c else "" for c in row]))
                    )

        for link_tab in link_tabs:
            link_headers = []
            for row in wb[link_tab].iter_rows(values_only=True):
                if not link_headers:
                    link_headers = [str(i).strip() for i in row]
                    translate_headers(link_headers, link_headers_map)
                else:
                    graph_dict["links"].append(
                        dict(zip(link_headers, [c if c else "" for c in row]))
                    )

        # add data to graph
        drawing.from_dict(graph_dict, diagram_name="Page-1")
    except:
        log.error("N2G:xlsx_data, failed: {}".format(traceback.format_exc()))
        return False

    # clean up
    wb.close()
    del graph_dict, wb

    return True
